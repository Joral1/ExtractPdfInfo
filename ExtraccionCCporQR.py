import os
import glob
from openpyxl import Workbook
from pdf2image import convert_from_path
from pdf2image.exceptions import PDFInfoNotInstalledError
import pytesseract
from tqdm import tqdm

# Ruta principal
#ruta_principal = r'C:\Users\USUARIO\Documents\MEL\Ataco\prueba'
#ruta_principal = r'C:\Users\USUARIO\Documents\MEL\Ataco\Prueba2'
#ruta_principal = r'C:\Users\USUARIO\Documents\MEL\Ataco\Prueba3'
#ruta_principal = r'C:\Users\USUARIO\Documents\MEL\Ataco\Prueba4'
ruta_principal = r'C:\Users\USUARIO\OneDrive - Tetra Tech, Inc\Documents\Ataco_Comprimido'
ruta_informe = r'C:\Users\USUARIO\Documents\MEL\Ataco\CC_Ataco.xlsx'

# Configuración de pytesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Ruta de Poppler
poppler_path = r'C:\poppler\Library\bin'  # Ruta correcta de Poppler

# Función para encontrar archivos PDF en una carpeta
def encontrar_archivos_pdf(ruta_principal):
    archivos_pdf = []
    for carpeta, _, _ in os.walk(ruta_principal):
        for archivo in glob.glob(os.path.join(carpeta, '*.pdf')):
            archivos_pdf.append(archivo)
    return archivos_pdf

# Función para extraer el contenido entre dos marcadores
def extraer_contenido_entre_marcadores(texto, marcador_inicio, marcador_fin):
    inicio_contenido = texto.lower().find(marcador_inicio.lower())
    fin_contenido = texto.lower().find(marcador_fin.lower(), inicio_contenido + len(marcador_inicio))
    if inicio_contenido != -1 and fin_contenido != -1:
        contenido = texto[inicio_contenido + len(marcador_inicio):fin_contenido]
        return contenido.strip()  # Eliminar espacios en blanco al inicio y al final
    else:
        print(f"No se encontraron los marcadores '{marcador_inicio}' y '{marcador_fin}' en el texto.")
        return ""

# Función para extraer texto de archivos PDF
def extraer_texto_pdf(archivo_pdf):
    texto = ""
    try:
        # Extraer texto del PDF
        imagenes = convert_from_path(archivo_pdf, first_page=1, last_page=10, poppler_path=poppler_path)
        for imagen in imagenes:
            texto += pytesseract.image_to_string(imagen)

        # Extraer contenido entre "ARTÍCULO SEGUNDO:" y "ARTÍCULO TERCERO:"
        contenido = extraer_contenido_entre_marcadores(texto, "ARTICULO SEGUNDO:", "ARTICULO TERCERO:")
        if contenido:
            return contenido

    except PDFInfoNotInstalledError:
        print("Poppler no está instalado o configurado correctamente.")
    except Exception as e:
        print(f"Error al extraer texto del archivo PDF {archivo_pdf}: {e}")
    return None

# Función para generar el archivo de Excel
def generar_excel(ruta_informe, lista_archivos_pdf):
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"
    ws['A1'] = 'Nombre del PDF'
    ws['B1'] = 'Información extraída'

    fila = 2
    total_archivos = len(lista_archivos_pdf)
    archivos_procesados = 0

    # Procesar los archivos en lotes de 20
    for i in range(0, total_archivos, 20):
        lotes_archivos_pdf = lista_archivos_pdf[i:i+20]
        with tqdm(total=len(lotes_archivos_pdf), desc="Procesando archivos PDF") as pbar:
            for archivo_pdf in lotes_archivos_pdf:
                nombre_pdf = os.path.basename(archivo_pdf)
                contenido = extraer_texto_pdf(archivo_pdf)
                if contenido is not None:  # Verificar si se encontró contenido
                    ws.cell(row=fila, column=1, value=nombre_pdf)  # Guardar el nombre del PDF en la columna A
                    ws.cell(row=fila, column=2, value=contenido)  # Guardar el contenido en la columna B
                    fila += 1
                pbar.update(1)
                archivos_procesados += 1

                # Actualizar el archivo de Excel después de procesar cada lote de 20 archivos
                if archivos_procesados % 20 == 0:
                    wb.save(ruta_informe)

    # Guardar el archivo de Excel final
    wb.save(ruta_informe)
    print(f"El informe se ha generado correctamente en {ruta_informe}")

# Obtener la lista de archivos PDF en la carpeta principal
archivos_pdf = encontrar_archivos_pdf(ruta_principal)

# Generar el archivo de Excel
generar_excel(ruta_informe, archivos_pdf)
